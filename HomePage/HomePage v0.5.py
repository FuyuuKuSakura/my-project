import os
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QAction, QLabel, QPushButton, QDialog, QVBoxLayout, QInputDialog, QLineEdit, QStatusBar
from PyQt5.QtGui import QPalette, QBrush, QPixmap
from PyQt5.QtCore import Qt, QTimer, QTime, QDate
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import QUrl
import random
import openpyxl
from openpyxl import Workbook, load_workbook
import requests as r



class MainWindow(QMainWindow):
    
    def __init__(self):
        super().__init__()
        self.initUI()
        self.initDateTime()  # 初始化日期时间显示

    def initDateTime(self):
        # 创建日期和时间标签
        self.date_label = QLabel(self)
        self.time_label = QLabel(self)
        
        # 样式设置
        label_style = """
            background-color: white;
            color: #b98a7c;
            font-size: 24px;
            padding: 5px;
            border-radius: 3px;
            font: Arial;
        """
        self.date_label.setStyleSheet(label_style)
        self.time_label.setStyleSheet(label_style)
        self.date_label.setAlignment(Qt.AlignCenter)
        self.time_label.setAlignment(Qt.AlignCenter)
        
        # 创建定时器更新时钟
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.updateDateTime)
        self.timer.start(1000)  # 每秒更新一次
        
        # 初始更新
        self.updateDateTime()
        self.updateLabelPositions()

    def updateDateTime(self):
        """更新日期和时间显示"""
        current_date = QDate.currentDate().toString('yyyy-MM-dd')
        current_time = QTime.currentTime().toString('hh:mm:ss')
        
        self.date_label.setText(current_date)
        self.time_label.setText(current_time)
        
        # 调整位置（确保在窗口调整大小时也能正确显示）
        self.updateLabelPositions()

    def updateLabelPositions(self):
        """更新标签位置到右上角"""
        label_width = 150
        label_height = 30
        margin = 0
        self.date_label.setGeometry(
            565, 
            margin , 
            label_width, 
            label_height
        )
        self.time_label.setGeometry(
            570 + 130 + 5,  # 在日期标签右侧，添加10px间距
            margin ,  # 添加10px间距
            label_width, 
            label_height
        )

    def resizeEvent(self, event):
        """窗口大小改变时的事件处理"""
        self.updateLabelPositions()
        self.update_background()
        super().resizeEvent(event)

    def initUI(self):
        # 设置窗口标题和初始大小
        self.setWindowTitle('My HomePage')
        self.setGeometry(300, 300, 1440, 797)

        # 创建菜单栏
        menubar = self.menuBar()
        fileMenu = menubar.addMenu('File')
        exitAction = QAction('Exit', self)
        exitAction.triggered.connect(self.close)
        fileMenu.addAction(exitAction)
        menubar.setNativeMenuBar(False)

        # 添加按钮
        self.createButtons()
        
        # 创建状态栏
        self.createStatusBar()
        
        # 设置背景图片
        self.setupBackground()


    def createButtons(self):
        """创建所有按钮"""
        self.button1 = QPushButton('吃什么好', self)
        self.button1.setGeometry(20, 110, 140, 50)
        self.button1.clicked.connect(self.open_random_dialog)
        self.button1.setVisible(False)
        
        self.button2 = QPushButton('打开课程网站', self)
        self.button2.setGeometry(20, 50, 200, 50)
        self.button2.clicked.connect(self.open_course_selection_dialog)
        self.button2.setVisible(False)

        
        self.button3 = QPushButton('todo', self)
        self.button3.setGeometry(20, 170, 80, 50)
        self.button3.clicked.connect(self.open_todo_list_dialog)
        self.button3.setVisible(False)


        button4 = QPushButton('·', self)
        button4.setGeometry(720, 800, 50, 50)
        button4.clicked.connect(self.display_applications)        

        self.button5 = QPushButton('译', self)
        self.button5.setGeometry(170, 110, 50, 100)
        self.button5.clicked.connect(self.open_translate_dialog)
        self.button5.setVisible(False)

        self.button6 = QPushButton('常用页面', self)
        self.button6.setGeometry(20, 290, 140, 50)
        self.button6.clicked.connect(self.useful_pages_dialog)


        # 设置按钮样式
        button_style = """
        QPushButton {
            background-color: #f9d6db;
            color: black;
            font-size: 16px;
            border: none;
            border-radius: 5px;
            padding: 10px 20px;
        }
        QPushButton:hover {
            background-color: #e8a3ae;
        }
        """
        
        button4_style = """
        QPushButton {
            background-color: rbga(255,255,255,0);
            color: white;
            font-size: 25px;
            border: none;
            border-radius: 25px;
            padding: 10px 20px;
        }
        QPushButton:hover {
            background-color: #f9d6db;
        }
        """
        
        self.button1.setStyleSheet(button_style)
        self.button2.setStyleSheet(button_style)
        self.button3.setStyleSheet(button_style)
        button4.setStyleSheet(button4_style)
        self.button5.setStyleSheet(button_style)

    def useful_pages_dialog(self):
        """打开常用页面对话框"""
        

    def display_applications(self) :
        if self.button1.isVisible():
            self.button1.setVisible(False)
        else:
            self.button1.setVisible(True)
        if self.button2.isVisible():
            self.button2.setVisible(False)
        else:
            self.button2.setVisible(True)
        if self.button3.isVisible():
            self.button3.setVisible(False)
        else:
            self.button3.setVisible(True)
        if self.button5.isVisible():
            self.button5.setVisible(False)
        else:
            self.button5.setVisible(True)

    def open_course_selection_dialog(self):
        """打开课程网站选择对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("请选择要打开的课程网站")
        dialog.setFixedSize(300, 300)  # 设置对话框大小
        layout = QVBoxLayout()

        courses = {
            "砺儒云": "https://moodle.scnu.edu.cn/login/index.php",
            "AbdnIMS": "https://abdn.blackboard.com/?new_loc=%2Fultra%2Fcourse",
            "iTest": "https://sso.unipus.cn/sso/login?service=https%3A%2F%2Fitestcloud.unipus.cn%2Futest%2Fitest%2Flogin%3F_rp%3D%252Fitest%253Fx%253D1760636722165",
            "WeLearn": "https://welearn.sflep.com/student/index.aspx",
            "超星学习通": "https://v8.chaoxing.com/"
        }

        def create_open_url_lambda(url):
            return lambda: (QDesktopServices.openUrl(QUrl(url)), dialog.accept())

        for name, url in courses.items():
            btn = QPushButton(name)
            btn.clicked.connect(create_open_url_lambda(url))
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #f9d6db;
                    padding: 10px;
                    margin: 2px;
                    border: 1px solid #ccc;
                    border-radius: 5px;
                    color: black;
                    font-size: 16px;
                }
                QPushButton:hover {
                    background-color: #e8a3ae;
                }
            """)
            layout.addWidget(btn)

        dialog.setLayout(layout)
        dialog.exec_()
    
    def open_random_dialog(self):
        """打开随机数对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("吃什么")
        dialog.setFixedSize(300, 300)  # 设置对话框大小
        layout = QVBoxLayout()

        # 创建文本输入框
        input_box = QLineEdit(dialog)
        input_box.setPlaceholderText("一个数喵")

        # 创建按钮
        button = QPushButton("生成随机数", dialog)
        button.setStyleSheet("""
                QPushButton {
                    background-color: #f9d6db;
                    padding: 10px;
                    margin: 2px;
                    border: 1px solid #ccc;
                    border-radius: 5px;
                    color: black;
                    font-size: 16px;
                }
                QPushButton:hover {
                    background-color: #e8a3ae;
                }
            """)

        # 创建用于显示结果的标签
        result_label = QLabel("", dialog)
        result_label.setAlignment(Qt.AlignCenter)


        def generate_random():
            max_val_text = input_box.text()
            try:
                max_val = int(max_val_text)
                if max_val > 0:
                    random_num = random.randint(1, max_val)
                    result_label.setText(f"喵喵: {random_num}")
                else:
                    result_label.setText("一个正整数喵")
            except ValueError:
                result_label.setText("好好吃饭喵。")

        button.clicked.connect(generate_random)

        # 将控件添加到布局中
        layout.addWidget(QLabel("一个数喵:"))
        layout.addWidget(input_box)
        layout.addWidget(button)
        layout.addWidget(result_label)

        dialog.setLayout(layout)
        dialog.exec_()
    
    def open_todo_list_dialog(self):
        """打开待办事项对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("待办事项")
        dialog.setFixedSize(300, 300)  # 设置对话框大小
        layout = QVBoxLayout()
        # 创建文本输入框
        input_event = QLineEdit(dialog)
        input_event.setPlaceholderText("输入事件")
        
        input_date = QLineEdit(dialog)
        input_date.setPlaceholderText("输入日期 (YYYY-MM-DD)")

        add_button = QPushButton("添加待办", dialog)
        
        # Excel文件操作
        excel_file = 'to_do_list.xlsx'

        def init_excel():
            if not os.path.exists(excel_file):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Tasks"
                sheet.append(["Event", "Date"])  # 添加表头
                workbook.save(excel_file)

        def add_task():
            event = input_event.text()
            date = input_date.text()
            if event and date:
                try:
                    workbook = load_workbook(excel_file)
                    sheet = workbook.active
                    sheet.append([event, date])
                    workbook.save(excel_file)
                    
                    input_event.clear()
                    input_date.clear()
                    self.statusBar().showMessage("待办事项已添加", 2000) # 显示2秒
                except Exception as e:
                    self.statusBar().showMessage(f"文件错误: {e}", 2000)
            else:
                self.statusBar().showMessage("事件和日期不能为空", 2000)

        init_excel() # 确保Excel文件和表已创建
        add_button.clicked.connect(add_task)

        # 布局
        layout.addWidget(QLabel("事件:"))
        layout.addWidget(input_event)
        layout.addWidget(QLabel("日期:"))
        layout.addWidget(input_date)
        layout.addWidget(add_button)

        dialog.setLayout(layout)
        dialog.exec_()

    def open_translate_dialog(self):
        """打开翻译对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("翻译")
        dialog.setFixedSize(800, 600)  # 设置对话框大小
        layout = QVBoxLayout()

        # 创建文本输入框
        input_text = QLineEdit(dialog)
        input_text.setPlaceholderText("输入要翻译的文本")

        # 创建按钮
        translate_button = QPushButton('翻译', dialog)

        # 创建用于显示结果的标签
        result_label = QLabel("", dialog)
        result_label.setAlignment(Qt.AlignCenter)

        def perform_translation():
            query = input_text.text()
            if query:
                translated_text = self.translate_api(query)
                result_label.setText(translated_text)
            else:
                result_label.setText("请输入文本进行翻译")

        translate_button.clicked.connect(perform_translation)

        # 将控件添加到布局中
        layout.addWidget(QLabel("要翻译的文本:"))
        layout.addWidget(input_text)
        layout.addWidget(translate_button)
        layout.addWidget(result_label)

        dialog.setLayout(layout)
        dialog.exec_()

    def translate_api(self, query, from_lang='AUTO', to_lang='AUTO'):
        url = 'https://fanyi.baidu.com/sug'
        data = {'kw': query} # 你只需要改kw对应的值
        res = r.post(url, data=data).json()
        return res['data'][0]['v']

    def createStatusBar(self):
        """创建状态栏"""
        status_bar = QStatusBar()
        self.setStatusBar(status_bar)
        status_bar.setStyleSheet("""
            QStatusBar {
                background-color: rbga(249,214,219,30);
                color: white;
                qproperty-alignment: AlignCenter;
            }
            QStatusBar::item {
                border: none;
            }
        """)
        status_bar.showMessage("贴贴成功！", 5000)


    def setupBackground(self):
        """设置背景图片"""
        self.setWindowOpacity(0.8)
        # 搜索图片文件
        self.image_path = ""
        for root, dirs, files in os.walk(os.path.expanduser("~")):
            if "IMG_9881.JPG" in files:
                self.image_path = os.path.join(root, "IMG_9881.JPG")
                break
        
        self.update_background()

    def update_background(self):
        """更新背景图片"""
        if os.path.exists(self.image_path):
            palette = QPalette()
            pixmap = QPixmap(self.image_path)
            scaled_pixmap = pixmap.scaled(self.size(), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            palette.setBrush(QPalette.Window, QBrush(scaled_pixmap))
            self.setPalette(palette)
        else:
            print(f"Error: Image file not found at {self.image_path}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MainWindow()
    ex.show()
    ex.showFullScreen()
    sys.exit(app.exec_())